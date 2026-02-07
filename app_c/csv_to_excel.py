#!/usr/bin/env python3
"""
CSVファイル群を個人ごとにExcelファイルにまとめるスクリプト
box/sample.xlsx の形式に従って出力する
"""

import json
import os
import csv
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.datavalidation import DataValidation

# カテゴリ名とシート名のマッピング
CATEGORY_MAP = {
    '論文': '論文',
    '分担執筆': '分担執筆',
    '単著': '単著',
    '共著編著': '共著・編著',
    '口頭発表': '口頭発表',
    'MISC': 'MISC',
    'その他': 'その他',
}

# 色の定義
FILL_BLUE = PatternFill(start_color='FFDAE3F3', end_color='FFDAE3F3', fill_type='solid')
FILL_GRAY = PatternFill(start_color='FFD6DCE4', end_color='FFD6DCE4', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFFFFFCC', end_color='FFFFFFCC', fill_type='solid')

# フォント定義
FONT_BLACK = Font(color='FF000000')
FONT_RED = Font(color='FFFF0000')

# 罫線定義
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 中央揃え定義
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

# 各シートのヘッダー構造（サンプルから抽出）
# 形式: (列番号, 行2のヘッダー, 行3のサブヘッダー(あれば), 色, CSVカラム名)
SHEET_STRUCTURES = {
    '論文': {
        'title': '【論文】赤字：必須項目',
        'columns': [
            (1, '機構通し番号', None, 'blue', None),
            (2, '更新日時', None, 'blue', None),
            (3, 'No.', None, 'gray', 'No.'),
            (4, '入力者名', None, 'yellow', '入力者名'),
            (5, 'e-Rad研究者番号', None, 'yellow', 'e-Rad研究者番号'),
            (6, '共同研究番号', None, 'yellow', '共同研究番号'),
            (7, '科研費課題番号', None, 'yellow', '科研費課題番号'),
            (8, '著者氏名（共著者含）', '原文', 'yellow', '著者氏名（共著者含）　＞　原文'),
            (9, None, '英訳', 'yellow', '著者氏名（共著者含）　＞　英訳'),
            (10, None, '下線', 'yellow', '著者氏名（共著者含）　＞　下線'),
            (11, '論文題目名', '原文', 'yellow', '論文題目名　＞　原文'),
            (12, None, '英訳', 'yellow', '論文題目名　＞　英訳'),
            (13, '記述言語', None, 'yellow', '記述言語'),
            (14, '掲載種別', None, 'yellow', '掲載種別'),
            (15, '査読', None, 'yellow', '査読'),
            (16, '招待論文', None, 'yellow', '招待論文'),
            (17, '国際共著', None, 'yellow', '国際共著'),
            (18, '掲載誌名', '原文', 'yellow', '掲載誌名　＞　原文'),
            (19, None, '英訳', 'yellow', '掲載誌名　＞　英訳'),
            (20, '掲載誌(巻・号・頁)', '巻', 'yellow', '掲載誌(巻・号・頁)　＞　巻'),
            (21, None, '号', 'yellow', '掲載誌(巻・号・頁)　＞　号'),
            (22, None, '開始頁', 'yellow', '掲載誌(巻・号・頁)　＞　開始頁'),
            (23, None, '終了頁', 'yellow', '掲載誌(巻・号・頁)　＞　終了頁'),
            (24, '掲載誌 発行年月(日)', None, 'yellow', '掲載誌　発行年月(日)'),
            (25, '出版機関名', '原文', 'yellow', '出版機関名　＞　原文'),
            (26, None, '英訳', 'yellow', '出版機関名　＞　英訳'),
            (27, '出版機関の所在地', None, 'yellow', '出版機関の所在地'),
            (28, 'ISBN', None, 'yellow', 'ISBN'),
            (29, 'ISSN', None, 'yellow', 'ISSN'),
            (30, 'DOI', None, 'yellow', 'DOI'),
            (31, 'CiNiiのID', None, 'yellow', 'CiNiiのID'),
            (32, '共著区分', None, 'yellow', '共著区分'),
            (33, '共著範囲', None, 'yellow', '共著範囲'),
            (34, '参加形態', None, 'yellow', '参加形態'),
            (35, '担当部分', None, 'yellow', '担当部分'),
            (36, 'リンクURL', 'Permalink', 'yellow', 'リンクURL　＞　Permalink'),
            (37, None, 'URL', 'yellow', 'リンクURL　＞　URL'),
            (38, 'Web of ScienceのID', None, 'yellow', 'Web of ScienceのID'),
            (39, 'PubMedのID', None, 'yellow', 'PubMedのID'),
            (40, 'ScopusのID', None, 'yellow', 'ScopusのID'),
            (41, 'JGlobalのID', None, 'yellow', 'JGlobalのID'),
            (42, 'arXivのID', None, 'yellow', 'arXivのID'),
            (43, 'ORCIDのPut Code', None, 'yellow', 'ORCIDのPut Code'),
            (44, 'DBLPのID', None, 'yellow', 'DBLPのID'),
            (45, 'OpenDepoのID', None, 'yellow', 'OpenDepoのID'),
            (46, '概要', '原文', 'yellow', '概要　＞　原文'),
            (47, None, '英訳', 'yellow', '概要　＞　英訳'),
            (48, 'メモ', None, 'yellow', 'メモ'),
            (49, '指標用フラグ', None, 'gray', None),
            (50, '備考１（機関担当者記入欄）', None, 'gray', None),
            (51, '備考２（機関担当者記入欄）', None, 'gray', None),
            (52, '備考３（機関担当者記入欄）', None, 'gray', None),
        ],
        'date_columns': [24],
        'required_row2': [4, 5, 6, 8, 11, 13, 14, 15, 17, 18, 24, 27],
        'required_row3': [8, 11, 18],
    },
    '分担執筆': {
        'title': '【分担執筆】赤字：必須項目',
        'columns': [
            (1, '機構通し番号', None, 'blue', None),
            (2, '更新日時', None, 'blue', None),
            (3, 'No.', None, 'gray', 'No.'),
            (4, '入力者名', None, 'yellow', '入力者名'),
            (5, 'e-Rad研究者番号', None, 'yellow', 'e-Rad研究者番号'),
            (6, '共同研究番号', None, 'yellow', '共同研究番号'),
            (7, '科研費課題番号', None, 'yellow', '科研費課題番号'),
            (8, '著者氏名（共著者含）', '原文', 'yellow', '著者氏名（共著者含）　＞　原文'),
            (9, None, '英訳', 'yellow', '著者氏名（共著者含）　＞　英訳'),
            (10, None, '下線', 'yellow', '著者氏名（共著者含）　＞　下線'),
            (11, '担当部分', '原文', 'yellow', '担当部分　＞　原文'),
            (12, None, '英訳', 'yellow', '担当部分　＞　英訳'),
            (13, '著書名', '原文', 'yellow', '著書名　＞　原文'),
            (14, None, '英訳', 'yellow', '著書名　＞　英訳'),
            (15, '記述言語', None, 'yellow', '記述言語'),
            (16, '著書種別', None, 'yellow', '著書種別'),
            (17, '出版機関名', '原文', 'yellow', '出版機関名　＞　原文'),
            (18, None, '英訳', 'yellow', '出版機関名　＞　英訳'),
            (19, '出版機関の所在地', None, 'yellow', '出版機関の所在地'),
            (20, '発行年月(日)', None, 'yellow', '発行年月(日)'),
            (21, '査読', None, 'yellow', '査読'),
            (22, 'ISBN', None, 'yellow', 'ISBN'),
            (23, 'ISSN', None, 'yellow', 'ISSN'),
            (24, 'DOI', None, 'yellow', 'DOI'),
            (25, '編者名', None, 'yellow', '編者名'),
            (26, '担当ページ', None, 'yellow', '担当ページ'),
            (27, '担当部分のページ数', None, 'yellow', '担当部分のページ数'),
            (28, '担当部分の共著区分', None, 'yellow', '担当部分の共著区分'),
            (29, '担当部分の共著範囲', None, 'yellow', '担当部分の共著範囲'),
            (30, '国際共著', None, 'yellow', '国際共著'),
            (31, 'URL', None, 'yellow', 'URL'),
            (32, 'ASIN', None, 'yellow', 'ASIN'),
            (33, 'Amazon URL', None, 'yellow', 'Amazon　URL'),
            (34, '概要', '原文', 'yellow', '概要　＞　原文'),
            (35, None, '英訳', 'yellow', '概要　＞　英訳'),
            (36, 'メモ', None, 'yellow', 'メモ'),
            (37, '指標用フラグ', None, 'gray', None),
            (38, '備考１（機関担当者記入欄）', None, 'gray', None),
            (39, '備考２（機関担当者記入欄）', None, 'gray', None),
            (40, '備考３（機関担当者記入欄）', None, 'gray', None),
        ],
        'date_columns': [20],
        'required_row2': [4, 5, 6, 8, 11, 13, 15, 16, 17, 19, 20, 21],
        'required_row3': [8, 11, 13, 17],
    },
    '単著': {
        'title': '【単著】赤字：必須項目',
        'columns': [
            (1, '機構通し番号', None, 'blue', None),
            (2, '更新日時', None, 'blue', None),
            (3, 'No.', None, 'gray', 'No.'),
            (4, '入力者名', None, 'yellow', '入力者名'),
            (5, 'e-Rad研究者番号', None, 'yellow', 'e-Rad研究者番号'),
            (6, '共同研究番号', None, 'yellow', '共同研究番号'),
            (7, '科研費課題番号', None, 'yellow', '科研費課題番号'),
            (8, '著者氏名', '原文', 'yellow', '著者氏名　＞　原文'),
            (9, None, '英訳', 'yellow', '著者氏名　＞　英訳'),
            (10, '著書名', '原文', 'yellow', '著書名　＞　原文'),
            (11, None, '英訳', 'yellow', '著書名　＞　英訳'),
            (12, '記述言語', None, 'yellow', '記述言語'),
            (13, '著書種別', None, 'yellow', '著書種別'),
            (14, '出版機関名', '原文', 'yellow', '出版機関名　＞　原文'),
            (15, None, '英訳', 'yellow', '出版機関名　＞　英訳'),
            (16, '出版機関の所在地', None, 'yellow', '出版機関の所在地'),
            (17, '発行年月(日)', None, 'yellow', '発行年月(日)'),
            (18, 'ISBN', None, 'yellow', 'ISBN'),
            (19, 'ISSN', None, 'yellow', 'ISSN'),
            (20, 'DOI', None, 'yellow', 'DOI'),
            (21, '査読', None, 'yellow', '査読'),
            (22, '章数', None, 'yellow', '章数'),
            (23, 'ページ数', None, 'yellow', 'ページ数'),
            (24, 'URL', None, 'yellow', 'URL'),
            (25, 'ASIN', None, 'yellow', 'ASIN'),
            (26, 'Amazon URL', None, 'yellow', 'Amazon　URL'),
            (27, '概要', '原文', 'yellow', '概要　＞　原文'),
            (28, None, '英訳', 'yellow', '概要　＞　英訳'),
            (29, 'メモ', None, 'yellow', 'メモ'),
            (30, '備考１（機関担当者記入欄）', None, 'gray', None),
            (31, '備考２（機関担当者記入欄）', None, 'gray', None),
            (32, '備考３（機関担当者記入欄）', None, 'gray', None),
        ],
        'date_columns': [17],
        'required_row2': [4, 5, 6, 8, 10, 12, 13, 14, 16, 17, 18],
        'required_row3': [8, 10, 14],
    },
    '共著・編著': {
        'title': '【共著・編著】赤字：必須項目',
        'columns': [
            (1, '機構通し番号', None, 'blue', None),
            (2, '更新日時', None, 'blue', None),
            (3, 'No.', None, 'gray', 'No.'),
            (4, '入力者名', None, 'yellow', '入力者名'),
            (5, 'e-Rad研究者番号', None, 'yellow', 'e-Rad研究者番号'),
            (6, '共同研究番号', None, 'yellow', '共同研究番号'),
            (7, '科研費課題番号', None, 'yellow', '科研費課題番号'),
            (8, '著者氏名（共著者含）', '原文', 'yellow', '著者氏名（共著者含）　＞　原文'),
            (9, None, '英訳', 'yellow', '著者氏名（共著者含）　＞　英訳'),
            (10, None, '下線', 'yellow', '著者氏名（共著者含）　＞　下線'),
            (11, '著書名', '原文', 'yellow', '著書名　＞　原文'),
            (12, None, '英訳', 'yellow', '著書名　＞　英訳'),
            (13, '記述言語', None, 'yellow', '記述言語'),
            (14, '著書種別', None, 'yellow', '著書種別'),
            (15, '出版機関名', '原文', 'yellow', '出版機関名　＞　原文'),
            (16, None, '英訳', 'yellow', '出版機関名　＞　英訳'),
            (17, '出版機関の所在地', None, 'yellow', '出版機関の所在地'),
            (18, '発行年月(日)', None, 'yellow', '発行年月(日)'),
            (19, 'ISBN', None, 'yellow', 'ISBN'),
            (20, 'ISSN', None, 'yellow', 'ISSN'),
            (21, 'DOI', None, 'yellow', 'DOI'),
            (22, '査読', None, 'yellow', '査読'),
            (23, '著書全体のページ数', None, 'yellow', 'ページ数著書全体のページ数'),
            (24, '著書形態', None, 'yellow', '著書形態'),
            (25, '共著範囲', None, 'yellow', '共著範囲'),
            (26, '国際共著', None, 'yellow', '国際共著'),
            (27, '執筆形態', None, 'yellow', '執筆形態'),
            (28, '著者として関わった章数', None, 'yellow', '著者として関わった章数'),
            (29, '担当ページ', None, 'yellow', '担当ページ'),
            (30, '担当部分', '原文', 'yellow', '担当部分　＞　原文'),
            (31, None, '英訳', 'yellow', '担当部分　＞　英訳'),
            (32, 'URL', None, 'yellow', 'URL'),
            (33, 'ASIN', None, 'yellow', 'ASIN'),
            (34, 'Amazon URL', None, 'yellow', 'Amazon　URL'),
            (35, '概要', '原文', 'yellow', '概要　＞　原文'),
            (36, None, '英訳', 'yellow', '概要　＞　英訳'),
            (37, 'メモ', None, 'yellow', 'メモ'),
            (38, '備考１（機関担当者記入欄）', None, 'gray', None),
            (39, '備考２（機関担当者記入欄）', None, 'gray', None),
            (40, '備考３（機関担当者記入欄）', None, 'gray', None),
        ],
        'date_columns': [18],
        'required_row2': [4, 5, 6, 8, 11, 13, 14, 15, 17, 18, 19],
        'required_row3': [8, 11, 15],
    },
    '口頭発表': {
        'title': '【口頭発表】赤字：必須項目',
        'columns': [
            (1, '機構通し番号', None, 'blue', None),
            (2, '更新日時', None, 'blue', None),
            (3, 'No.', None, 'gray', 'No.'),
            (4, '入力者名', None, 'yellow', '入力者名'),
            (5, 'e-Rad研究者番号', None, 'yellow', 'e-Rad研究者番号'),
            (6, '共同研究番号', None, 'yellow', '共同研究番号'),
            (7, '科研費課題番号', None, 'yellow', '科研費課題番号'),
            (8, '発表者名（共同発表者含）', '原文', 'yellow', '発表者名（共同発表者含）　＞　原文'),
            (9, None, '英訳', 'yellow', '発表者名（共同発表者含）　＞　英訳'),
            (10, None, '下線', 'yellow', '発表者名（共同発表者含））　＞　下線'),
            (11, '題目又はセッション名', '原文', 'yellow', '題目又はセッション名　＞　原文'),
            (12, None, '英訳', 'yellow', '題目又はセッション名　＞　英訳'),
            (13, '会議区分', None, 'yellow', '会議区分'),
            (14, '会議名称', '原文', 'yellow', '会議名称　＞　原文'),
            (15, None, '英訳', 'yellow', '会議名称　＞　英訳'),
            (16, '主催者名称', '原文', 'yellow', '主催者名称 　＞　原文'),
            (17, None, '英訳', 'yellow', '主催者名称 　＞　英訳'),
            (18, '開催場所', '原文', 'yellow', '開催場所　＞　原文'),
            (19, None, '英訳', 'yellow', '開催場所　＞　英訳'),
            (20, None, '開催国', 'yellow', '開催場所　＞　開催国'),
            (21, '発表年月(日)', '（自）', 'yellow', '発表年月(日) 　＞　（自）'),
            (22, None, '（至）', 'yellow', '発表年月(日) 　＞　（至）'),
            (23, '発表形態', None, 'yellow', '発表形態'),
            (24, '発表(記述)言語', None, 'yellow', '発表(記述)言語'),
            (25, '招待の有無', None, 'yellow', '招待の有無'),
            (26, '査読', None, 'yellow', '査読'),
            (27, '共同作業範囲', None, 'yellow', '共同作業範囲'),
            (28, '担当部分', None, 'yellow', '担当部分'),
            (29, 'URL', None, 'yellow', 'URL'),
            (30, 'CiNiiのID', None, 'yellow', 'CiNiiのID'),
            (31, 'OpenDepoのID', None, 'yellow', 'OpenDepoのID'),
            (32, 'DOI', None, 'yellow', 'DOI'),
            (33, '概要', '原文', 'yellow', '概要　＞　原文'),
            (34, None, '英訳', 'yellow', '概要　＞　英訳'),
            (35, 'メモ', None, 'yellow', 'メモ'),
            (36, '指標用フラグ', None, 'gray', None),
            (37, '備考１（機関担当者記入欄）', None, 'gray', None),
            (38, '備考２（機関担当者記入欄）', None, 'gray', None),
            (39, '備考３（機関担当者記入欄）', None, 'gray', None),
        ],
        'date_columns': [21, 22],
        'required_row2': [4, 5, 6, 8, 11, 13, 14, 18, 21, 23, 24, 25],
        'required_row3': [8, 11, 14, 20, 21, 22],
    },
    'MISC': {
        'title': '【MISC】赤字：必須項目',
        'columns': [
            (1, '機構通し番号', None, 'blue', None),
            (2, '更新日時', None, 'blue', None),
            (3, 'No.', None, 'gray', 'No.'),
            (4, '入力者名', None, 'yellow', '入力者名'),
            (5, 'e-Rad研究者番号', None, 'yellow', 'e-Rad研究者番号'),
            (6, '共同研究番号', None, 'yellow', '共同研究番号'),
            (7, '科研費課題番号', None, 'yellow', '科研費課題番号'),
            (8, '著者氏名（共著者含）', '原文', 'yellow', '著者氏名（共著者含）　＞　原文'),
            (9, None, '英訳', 'yellow', '著者氏名（共著者含）　＞　英訳'),
            (10, None, '下線', 'yellow', '著者氏名（共著者含）　＞　下線'),
            (11, '題目', '原文', 'yellow', '題目　＞　原文'),
            (12, None, '英訳', 'yellow', '題目　＞　英訳'),
            (13, '記述言語', None, 'yellow', '記述言語'),
            (14, '掲載種別', None, 'yellow', '掲載種別'),
            (15, '掲載誌名', '原文', 'yellow', '掲載誌名　＞　原文'),
            (16, None, '英訳', 'yellow', '掲載誌名　＞　英訳'),
            (17, '掲載誌(巻・号・頁)', '巻', 'yellow', '掲載誌(巻・号・頁)　＞　巻'),
            (18, None, '号', 'yellow', '掲載誌(巻・号・頁)　＞　号'),
            (19, None, '開始頁', 'yellow', '掲載誌(巻・号・頁)　＞　開始頁'),
            (20, None, '終了頁', 'yellow', '掲載誌(巻・号・頁)　＞　終了頁'),
            (21, '掲載誌 発行年月(日)', None, 'yellow', '掲載誌　発行年月(日)'),
            (22, '出版機関名', '原文', 'yellow', '出版機関名　＞　原文'),
            (23, None, '英訳', 'yellow', '出版機関名　＞　英訳'),
            (24, '出版機関の所在地', None, 'yellow', '出版機関の所在地'),
            (25, '査読', None, 'yellow', '査読'),
            (26, '依頼の有無', None, 'yellow', '依頼の有無'),
            (27, 'ISBN', None, 'yellow', 'ISBN'),
            (28, 'ISSN', None, 'yellow', 'ISSN'),
            (29, 'DOI', None, 'yellow', 'DOI'),
            (30, 'CiNiiのID', None, 'yellow', 'CiNiiのID'),
            (31, '共著区分', None, 'yellow', '共著区分'),
            (32, '共著範囲', None, 'yellow', '共著範囲'),
            (33, '担当部分', None, 'yellow', '担当部分'),
            (34, 'リンクURL', 'Permalink', 'yellow', 'リンクURL　＞　Permalink'),
            (35, None, 'URL', 'yellow', 'リンクURL　＞　URL'),
            (36, 'PubMedのID', None, 'yellow', 'PubMedのID'),
            (37, 'JglobalのID', None, 'yellow', 'JGlobalのID'),
            (38, 'arXivのID', None, 'yellow', 'arXivのID'),
            (39, 'ORCIDのPut Code', None, 'yellow', 'ORCIDのPut Code'),
            (40, 'DBLPのID', None, 'yellow', 'DBLPのID'),
            (41, 'OpenDepoのID', None, 'yellow', 'OpenDepoのID'),
            (42, '概要', '原文', 'yellow', '概要　＞　原文'),
            (43, None, '英訳', 'yellow', '概要　＞　英訳'),
            (44, 'メモ', None, 'yellow', 'メモ'),
            (45, '備考１（機関担当者記入欄）', None, 'gray', None),
            (46, '備考２（機関担当者記入欄）', None, 'gray', None),
            (47, '備考３（機関担当者記入欄）', None, 'gray', None),
        ],
        'date_columns': [21],
        'required_row2': [4, 5, 6, 8, 11, 13, 14, 15, 21, 24],
        'required_row3': [8, 11, 15],
    },
    'その他': {
        'title': '【その他】赤字：必須項目',
        'columns': [
            (1, '機構通し番号', None, 'blue', None),
            (2, '更新日時', None, 'blue', None),
            (3, 'No.', None, 'gray', 'No.'),
            (4, '入力者名', None, 'yellow', '入力者名'),
            (5, 'e-Rad研究者番号', None, 'yellow', 'e-Rad研究者番号'),
            (6, '共同研究番号', None, 'yellow', '共同研究番号'),
            (7, '科研費課題番号', None, 'yellow', '科研費課題番号'),
            (8, '種別', None, 'yellow', '種別'),
            (9, '概要', None, 'yellow', '概要'),
            (10, '年月(日)', None, 'yellow', '年月(日)'),
            (11, 'メモ', None, 'yellow', 'メモ'),
            (12, '備考１（機関担当者記入欄）', None, 'gray', None),
            (13, '備考２（機関担当者記入欄）', None, 'gray', None),
            (14, '備考３（機関担当者記入欄）', None, 'gray', None),
        ],
        'date_columns': [10],
        'required_row2': [4, 5, 6, 8, 9, 10],
        'required_row3': [],
    },
}

# データ入力規則（ドロップダウンリスト）
# restrictions.json から読み込み
def load_data_validations():
    """restrictions.json からデータ入力規則を読み込む"""
    json_path = Path(__file__).parent.parent / "data" / "restrictions.json"
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    sep = data["_format"]["separator"]
    sep_nospace = data["_format"]["separator_nospace"]

    validations = {}
    for sheet_name, columns in data.items():
        if sheet_name.startswith("_"):
            continue

        # このシートで使用するセパレータ
        use_nospace = columns.get("_use_separator_nospace", False)
        current_sep = sep_nospace if use_nospace else sep

        sheet_validations = {}
        for key, choices in columns.items():
            if key.startswith("_"):
                continue
            # キー形式: "14_掲載種別" → 列番号 14
            col_num = int(key.split("_")[0])
            # 選択肢を "code : label" 形式の文字列に変換
            choice_strs = [f"{code}{current_sep}{label}" for code, label in choices]
            sheet_validations[col_num] = ",".join(choice_strs)

        validations[sheet_name] = sheet_validations

    return validations


DATA_VALIDATIONS = load_data_validations()


def get_fill(color_name):
    """色名からFillオブジェクトを取得"""
    if color_name == 'blue':
        return FILL_BLUE
    elif color_name == 'gray':
        return FILL_GRAY
    elif color_name == 'yellow':
        return FILL_YELLOW
    return None


def parse_date_value(value):
    """日付値を整数形式に変換（YYYYMMDD）"""
    if not value:
        return None
    # 文字列から数字以外を除去
    value_str = str(value).strip()
    # すでに数字のみなら整数に変換
    if value_str.isdigit():
        return int(value_str)
    # YYYY-MM-DD形式等の場合
    digits = re.sub(r'\D', '', value_str)
    if digits and len(digits) >= 6:
        return int(digits[:8]) if len(digits) >= 8 else int(digits)
    return value


def get_fiscal_year_range(fiscal_year):
    """年度の開始日・終了日を返す（YYYYMMDD整数）

    Args:
        fiscal_year: 年度（例: 2023 → 2023/4/1〜2024/3/31）

    Returns:
        (start_date, end_date) タプル、または None（フィルタなしの場合）
    """
    if fiscal_year is None:
        return None
    fy = int(fiscal_year)
    start_date = fy * 10000 + 401      # 20230401
    end_date = (fy + 1) * 10000 + 331  # 20240331
    return (start_date, end_date)


def is_in_fiscal_year(date_value, fiscal_year_range):
    """日付が年度範囲内かどうかを判定

    Args:
        date_value: YYYYMMDD形式の日付（整数または文字列）
        fiscal_year_range: (start_date, end_date) タプル

    Returns:
        True: 範囲内、False: 範囲外、None: 日付が不明
    """
    if fiscal_year_range is None:
        return True  # フィルタなし

    if not date_value:
        return None  # 日付不明

    # 整数に変換
    date_int = parse_date_value(date_value)
    if date_int is None or not isinstance(date_int, int):
        return None

    start_date, end_date = fiscal_year_range
    return start_date <= date_int <= end_date


def get_date_csv_columns(structure):
    """シート構造から日付列のCSVカラム名を取得"""
    date_col_nums = set(structure.get('date_columns', []))
    csv_col_names = []
    for col_num, header2, header3, color, csv_col in structure['columns']:
        if col_num in date_col_nums and csv_col:
            csv_col_names.append(csv_col)
    return csv_col_names


def read_csv_file(filepath, fiscal_year_range=None, date_csv_columns=None):
    """CSVファイルを読み込んでデータを返す

    Args:
        filepath: CSVファイルパス
        fiscal_year_range: 年度範囲 (start_date, end_date) または None
        date_csv_columns: フィルタ対象の日付列名リスト

    Returns:
        フィルタ済みのデータリスト
    """
    data = []
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 空行をスキップ
            if not any(v.strip() for v in row.values() if v):
                continue

            # 年度フィルタ
            if fiscal_year_range and date_csv_columns:
                # 日付列から最も適切な値を取得（複数ある場合は最初のものを使用）
                date_value = None
                for col_name in date_csv_columns:
                    if col_name in row and row[col_name]:
                        # 範囲形式（例: "20230401-20240331"）の場合は開始日を使用
                        val = row[col_name].split('-')[0].strip()
                        if val and val != '現在':
                            date_value = val
                            break

                in_range = is_in_fiscal_year(date_value, fiscal_year_range)
                # None（日付不明）の場合は含める、False（範囲外）の場合は除外
                if in_range is False:
                    continue

            data.append(row)
    return data


def create_sheet(wb, sheet_name, structure, csv_data=None):
    """シートを作成してフォーマットを適用"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    columns = structure['columns']
    date_columns = structure.get('date_columns', [])
    required_row2 = set(structure.get('required_row2', []))
    required_row3 = set(structure.get('required_row3', []))

    # 最大列番号を取得
    max_col = max(col[0] for col in columns)

    # 行1: タイトル（リッチテキストで「赤字：必須項目」を赤くする）
    title = structure['title']
    # タイトルを「【XXX】」と「赤字：必須項目」に分割
    if '赤字：必須項目' in title:
        prefix = title.replace('赤字：必須項目', '')
        red_font = InlineFont(color='FFFF0000')
        rich_title = CellRichText([prefix, TextBlock(red_font, '赤字：必須項目')])
        ws.cell(row=1, column=3, value=rich_title)
    else:
        ws.cell(row=1, column=3, value=title)

    # 行2-3: ヘッダーとセル結合
    # 結合情報を収集
    header_groups = []  # (start_col, end_col) for horizontal merge in row 2
    single_cols = []     # columns to merge vertically (row 2-3)

    current_group_start = None
    for i, (col_num, header2, header3, color, csv_col) in enumerate(columns):
        fill = get_fill(color)

        # 行2のヘッダー
        if header2:
            # 前のグループを閉じる
            if current_group_start is not None:
                prev_col = columns[i-1][0]
                if prev_col > current_group_start:
                    header_groups.append((current_group_start, prev_col))
                current_group_start = None

            # フォント色（必須項目は赤）
            font = FONT_RED if col_num in required_row2 else FONT_BLACK

            cell = ws.cell(row=2, column=col_num, value=header2)
            cell.font = font
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            if fill:
                cell.fill = fill

            # サブヘッダーがあるかチェック
            if header3:
                current_group_start = col_num
            else:
                # サブヘッダーがない場合は縦結合候補
                single_cols.append(col_num)
        else:
            # header2がNone = サブ項目の続き
            cell = ws.cell(row=2, column=col_num)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            if fill:
                cell.fill = fill

        # 行3のサブヘッダー
        if header3:
            # フォント色（必須項目は赤）
            font = FONT_RED if col_num in required_row3 else FONT_BLACK

            cell = ws.cell(row=3, column=col_num, value=header3)
            cell.font = font
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            if fill:
                cell.fill = fill
        else:
            cell = ws.cell(row=3, column=col_num)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            if fill:
                cell.fill = fill

    # 最後のグループを閉じる
    if current_group_start is not None:
        last_col = columns[-1][0]
        if last_col > current_group_start:
            header_groups.append((current_group_start, last_col))

    # セル結合を適用
    # 縦結合（サブヘッダーがない列）
    for col in single_cols:
        try:
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        except:
            pass  # すでに結合されている場合はスキップ

    # 横結合（サブヘッダーがある列グループ）
    for start_col, end_col in header_groups:
        try:
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        except:
            pass

    # CSVデータを挿入（行4から）
    if csv_data:
        # CSVのカラム名とExcel列の対応を作成
        col_mapping = {}
        for col_num, header2, header3, color, csv_col in columns:
            if csv_col:
                col_mapping[csv_col] = col_num

        for row_idx, row_data in enumerate(csv_data, start=4):
            for csv_col, col_num in col_mapping.items():
                if csv_col in row_data:
                    value = row_data[csv_col]
                    if value:
                        # 日付列は整数に変換
                        if col_num in date_columns:
                            value = parse_date_value(value)
                        # 数値に変換可能なら数値として書き込む
                        elif value.isdigit():
                            value = int(value)
                        cell = ws.cell(row=row_idx, column=col_num, value=value)
                        cell.border = THIN_BORDER

            # データがない列にも罫線を適用
            for col_num, _, _, _, _ in columns:
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = THIN_BORDER

    # データ入力規則（ドロップダウンリスト）を適用
    if sheet_name in DATA_VALIDATIONS:
        # データ行の範囲を決定（4行目から、データがあれば最終行+余裕、なければ100行程度）
        data_rows = len(csv_data) if csv_data else 0
        end_row = max(data_rows + 4 + 50, 100)  # データ行 + 余裕

        for col_num, formula in DATA_VALIDATIONS[sheet_name].items():
            col_letter = get_column_letter(col_num)
            cell_range = f"{col_letter}4:{col_letter}{end_row}"

            dv = DataValidation(
                type="list",
                formula1=f'"{formula}"',
                allow_blank=True,
                showDropDown=False,  # False = ドロップダウンを表示
            )
            dv.add(cell_range)
            ws.add_data_validation(dv)

    return ws


def get_researcher_files(csv_dir):
    """CSVディレクトリから研究者ごとのファイルを整理"""
    researchers = defaultdict(dict)

    for filepath in Path(csv_dir).glob('*.csv'):
        filename = filepath.stem
        # ファイル名形式: {researcher_id}-{category}.csv
        match = re.match(r'^(.+)-(.+)$', filename)
        if match:
            researcher_id = match.group(1)
            category = match.group(2)
            researchers[researcher_id][category] = filepath

    return researchers


def create_researcher_excel(researcher_id, category_files, output_dir, fiscal_year=None):
    """研究者のExcelファイルを作成

    Args:
        researcher_id: 研究者ID
        category_files: カテゴリ名 -> CSVファイルパスの辞書
        output_dir: 出力ディレクトリ
        fiscal_year: 年度フィルタ（例: 2023 = 2023年4月1日〜2024年3月31日）、Noneでフィルタなし
    """
    wb = Workbook()
    # デフォルトシートを削除
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 年度範囲を計算
    fiscal_year_range = get_fiscal_year_range(fiscal_year)

    # シート順序を保持
    sheet_order = ['論文', '分担執筆', '単著', '共著・編著', '口頭発表', 'MISC', 'その他']

    for sheet_name in sheet_order:
        # CSVカテゴリ名に変換
        csv_category = None
        for csv_cat, excel_sheet in CATEGORY_MAP.items():
            if excel_sheet == sheet_name:
                csv_category = csv_cat
                break

        # シート構造を取得
        structure = SHEET_STRUCTURES.get(sheet_name)
        if not structure:
            continue

        # CSVデータを読み込む（年度フィルタ適用）
        csv_data = None
        if csv_category and csv_category in category_files:
            date_csv_columns = get_date_csv_columns(structure)
            csv_data = read_csv_file(
                category_files[csv_category],
                fiscal_year_range=fiscal_year_range,
                date_csv_columns=date_csv_columns
            )

        # シートを作成
        create_sheet(wb, sheet_name, structure, csv_data)

    # ファイルを保存
    # 年度フィルタがある場合はファイル名に年度を付加
    if fiscal_year:
        output_path = Path(output_dir) / f'{researcher_id}_{fiscal_year}年度.xlsx'
    else:
        output_path = Path(output_dir) / f'{researcher_id}.xlsx'
    wb.save(output_path)
    print(f'Created: {output_path}')
    return output_path


def main():
    import argparse

    parser = argparse.ArgumentParser(description='CSVファイル群をExcelファイルに変換')
    parser.add_argument('--csv-dir', help='CSVファイルのディレクトリ')
    parser.add_argument('--output-dir', help='出力ディレクトリ')
    parser.add_argument('--fiscal-year', type=int, help='年度フィルタ（例: 2023 = 2023年4月〜2024年3月）')
    args = parser.parse_args()

    # パス設定
    base_dir = Path(__file__).parent.parent.parent
    csv_dir = Path(args.csv_dir) if args.csv_dir else base_dir / 'data' / 'csv'
    output_dir = Path(args.output_dir) if args.output_dir else base_dir / 'data' / 'xlsx'
    fiscal_year = args.fiscal_year

    # 出力ディレクトリ作成
    output_dir.mkdir(exist_ok=True)

    # 研究者ごとのファイルを取得
    researchers = get_researcher_files(csv_dir)

    if fiscal_year:
        print(f'Fiscal year filter: {fiscal_year}年度 ({fiscal_year}/4/1 - {fiscal_year + 1}/3/31)')

    print(f'Found {len(researchers)} researchers:')
    for researcher_id in researchers:
        print(f'  - {researcher_id}: {list(researchers[researcher_id].keys())}')

    # 各研究者のExcelファイルを作成
    for researcher_id, category_files in researchers.items():
        create_researcher_excel(researcher_id, category_files, output_dir, fiscal_year=fiscal_year)

    print(f'\nAll files created in: {output_dir}')


if __name__ == '__main__':
    main()
